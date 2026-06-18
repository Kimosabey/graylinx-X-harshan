#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_tracker.py — Graylinx x Harshan Work Tracker generator.

Source of truth for the tracker. It:
  1. Reads the curated project catalogue below.
  2. Fills real dates/commit counts from git history (where a repo exists),
     else estimates start/end from filesystem mtimes (flagged 'approx').
  3. Parses the deployment/attendance CSV into a day-level engagement log,
     carries it forward to TODAY (deriving days from git activity, flagging the
     rest 'needs-fill'), and cross-links each day to a project.
  4. Writes works.csv, engagement.csv, works.json and a styled works.xlsx.
  5. Injects the combined dataset as seed JSON into index.html.

Re-run any time to refresh from disk:  python build_tracker.py
The original attendance CSV is treated as read-only input.
"""

import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
from datetime import date, datetime, timedelta

# --------------------------------------------------------------------------- #
#  Paths / constants
# --------------------------------------------------------------------------- #
OUT_DIR = os.path.dirname(os.path.abspath(__file__))   # D:\Harshan\Graylinx X Harshan
BASE = os.path.dirname(OUT_DIR)                          # D:\Harshan
TODAY = date(2026, 6, 17)                                # "till now" anchor

ATTENDANCE_CSV = os.path.join(
    OUT_DIR, "data", "Harshan X Graylinx Deployment Tracker 2026(Deployment Attendance).csv"
)
INDEX_HTML = os.path.join(OUT_DIR, "index.html")

# --- Teams 1:1 chat exports -> per-day "communications" (topics, never verbatim) ----------
TEAMS_CHAT_DIR = os.path.join(OUT_DIR, "gl_teams_chat")
# file stem -> display name (in-file sender labels are unreliable; e.g. anushri's file says "Intern")
TEAMS_PEOPLE = {
    "vishnu": "Vishnu Ranga", "pradeep": "Pradeep Nagabhushana", "anushri_intern": "Anushri",
    "himanshu": "Himanshu", "karthik": "Karthik", "krishnaprasad": "Krishnaprasad",
    "niralipatel": "Nirali Patel", "santosh": "Santosh",
}
# What lands in the PUBLIC seed: "counts" (who+how-many) | "topics" (default, derived themes,
# no verbatim) | "full" (adds a short snippet — LOCAL builds only, never deployed).
TEAMS_DETAIL = os.environ.get("GLX_TEAMS_DETAIL", "topics")

# Monochrome black-and-white palette (neutral/achromatic) used for xlsx styling
C = {
    "ink": "171717", "ink2": "4B4B4B", "line": "C6C6C6",
    "base": "ECECEC", "card": "F4F4F4", "white": "FFFFFF",
    # status — neutral prominence ramp (Active darkest → Paused lightest)
    "Active": "171717", "Ongoing": "3A3A3A", "Done": "5C5C5C",
    "Planned": "888888", "Paused": "888888", "At-risk": "171717",
    # work types (neutral)
    "Office": "171717", "Discussion": "5C5C5C", "WFH": "888888", "Leave": "ABABAB",
}

# --------------------------------------------------------------------------- #
#  Curated project catalogue (static knowledge — dates filled at build time)
# --------------------------------------------------------------------------- #
# dir       : folder name under D:\Harshan (None = no folder)
# git       : True -> read dates/commits from git; else estimate from mtimes
# override  : {"start","end","source"} to hard-set dates (placeholders/empties)
# milestones: [{"name","status","progress","desc"}]
PROJECTS = [
    {
        "id": "thermynx", "name": "THERMYNX — HVAC AI Operations Intelligence",
        "dir": "HVAC AI Operations Intelligence Platform", "git": True,
        "category": "HVAC Platform", "status": "Active", "progress": 70,
        "description": "AI-powered HVAC analytics POC for the Unicharm facility — local LLM, RAG, work orders and analytics.",
        "role": "Sole engineer — full-stack build: FastAPI backend, React/TypeScript frontend, Ollama integration, analytics engine.",
        "tech": ["Python", "FastAPI", "React", "TypeScript", "Ollama", "MySQL", "Vite"],
        "highlights": [
            "Most active project — 308 commits",
            "18 feature pages + 6 Claude/LLM agents",
            "RAG over facility docs, work-order workflow, analytics engine",
            "Local on-prem LLM (Ollama) inference",
        ],
        "milestones": [
            {"name": "18 feature pages", "status": "Active", "progress": 80, "desc": "Frontend feature surface"},
            {"name": "6 AI agents", "status": "Active", "progress": 70, "desc": "Claude/LLM agents"},
            {"name": "RAG pipeline", "status": "Active", "progress": 65, "desc": "Retrieval over facility docs"},
            {"name": "Work orders", "status": "Active", "progress": 60, "desc": "Work-order workflow"},
            {"name": "Analytics engine", "status": "Active", "progress": 60, "desc": "Chiller/HVAC analytics"},
        ],
    },
    {
        "id": "climastream", "name": "ClimaStream — Real-time HVAC Pipeline",
        "dir": "clima-stream-hvac", "git": True,
        "category": "HVAC Platform", "status": "Done", "progress": 100,
        "description": "Production-grade streaming pipeline POC: Kafka -> TimescaleDB -> Airflow -> FastAPI -> Next.js.",
        "role": "Sole engineer — designed and built the end-to-end microservice pipeline and simulator.",
        "tech": ["Python", "Kafka", "TimescaleDB", "Airflow", "Next.js", "Docker"],
        "highlights": [
            "Proved the real-time streaming architecture",
            "Microservices: simulator, ingestion, processor, ML, storage, API",
            "31 commits over ~4 weeks",
        ],
        "milestones": [
            {"name": "Device simulator", "status": "Done", "progress": 100, "desc": "Synthetic HVAC data"},
            {"name": "Kafka ingestion", "status": "Done", "progress": 100, "desc": "Streaming ingest"},
            {"name": "TimescaleDB storage", "status": "Done", "progress": 100, "desc": "Time-series store"},
            {"name": "ML scoring", "status": "Done", "progress": 100, "desc": "Anomaly/score service"},
            {"name": "Next.js UI", "status": "Done", "progress": 100, "desc": "Frontend dashboard"},
        ],
    },
    {
        "id": "omnyx", "name": "OMNYX — Universal IoT Operations Platform",
        "dir": "graylinx-v2", "git": True,
        "category": "HVAC Platform", "status": "Active", "progress": 35,
        "description": "Next-gen production platform vision: Digital-Twin FDD, RL optimization, Agentic AI, dual-database architecture.",
        "role": "Architecture & platform design — the strategic north star consolidating the POCs.",
        "tech": ["Python", "React", "TypeScript", "PostgreSQL", "TimescaleDB", "Kafka"],
        "highlights": [
            "Digital Twin fault detection & diagnostics (FDD)",
            "Reinforcement-learning optimization",
            "Agentic AI operations layer",
            "Dual-DB (PostgreSQL + TimescaleDB) design",
        ],
        "milestones": [
            {"name": "Digital-Twin FDD", "status": "Active", "progress": 40, "desc": "Fault detection & diagnostics"},
            {"name": "RL optimization", "status": "Planned", "progress": 15, "desc": "Setpoint optimization"},
            {"name": "Agentic AI", "status": "Active", "progress": 30, "desc": "Operations agents"},
            {"name": "Dual-DB architecture", "status": "Active", "progress": 45, "desc": "Postgres + Timescale"},
        ],
    },
    {
        "id": "gl-pulse", "name": "GL Pulse — Platform Architecture",
        "dir": "gl-pulse", "git": False,
        "category": "HVAC Platform", "status": "Ongoing", "progress": 60,
        "description": "Reference architecture & planning docs unifying ClimaStream, THERMYNX and OMNYX into one next-gen vision.",
        "role": "Author — architecture documentation and lessons-learned consolidation.",
        "tech": ["Markdown", "Architecture"],
        "highlights": ["20 architecture documents", "Consolidates POC lessons into unified design"],
        "milestones": [],
    },
    {
        "id": "gl-pbs", "name": "GL PBS / BACnet Simulation",
        "dir": "gl-pbs-simulation", "git": False,
        "category": "HVAC Platform", "status": "Paused", "progress": 20,
        "description": "BACnet device simulator workstream (simulator dashboard lives under Current Works in GL).",
        "role": "Simulation flow study and BACnet device/UI analysis.",
        "tech": ["Vue 3", "BACnet", "Simulation"],
        "highlights": ["BACnet device simulation", "Vue 3 simulator dashboard"],
        "override": {"start": "2026-05-11", "end": "2026-05-13", "source": "manual"},
        "milestones": [],
    },
    {
        "id": "current-works", "name": "Current Works in GL — Core Monorepo",
        "dir": "Current Works in GL", "git": True,
        "category": "Graylinx Core", "status": "Active", "progress": 75,
        "description": "Active Graylinx monorepo: backend/frontend, commission tool, ML trainings, simulations, Kafka stress testing.",
        "role": "Sole engineer — Kafka performance testing, cross-machine benchmarking, DB/ETL architecture audits.",
        "tech": ["Node.js", "Express", "Kafka", "MySQL", "Python"],
        "highlights": [
            "Kafka stress tests 150x -> 5000x scale",
            "Cross-machine 24GB vs 8GB comparison reports",
            "DB & ETL critical-flaws audits, management summaries",
        ],
        "milestones": [
            {"name": "Kafka stress tests (150x-5000x)", "status": "Active", "progress": 80, "desc": "Throughput scaling"},
            {"name": "Cross-machine 24GB vs 8GB", "status": "Done", "progress": 100, "desc": "Hardware comparison"},
            {"name": "DB/ETL architecture audits", "status": "Active", "progress": 70, "desc": "Critical-flaws reviews"},
        ],
    },
    {
        "id": "docs", "name": "Graylinx Docs & PRD",
        "dir": "Docs", "git": False,
        "category": "Graylinx Core", "status": "Ongoing", "progress": 65,
        "description": "Product/architecture documentation: PRD, Modules & Features (175 features), Kafka/stress reports.",
        "role": "Author/maintainer — PRD, feature inventory, architecture references.",
        "tech": ["Markdown", "PRD", "Documentation"],
        "highlights": ["Graylinx_Modules_and_Features.md — 175 features, phased", "PRD v1.0 + Industrial Platform Master"],
        "milestones": [],
    },
    {
        "id": "selfaware-suite", "name": "The SelfAware® Intelligence Suite",
        "dir": "The SelfAware® Intelligence Suite", "git": True,
        "category": "SelfAware Suite", "status": "Active", "progress": 45,
        "description": "Six-product local-AI suite (6 repos): Continuum (HVAC intelligence), AuditShield, Autonomous-Cortex, NeuralPulse, VisioScan, SpatialNexus.",
        "role": "Designed & built the multi-product suite across six repositories.",
        "tech": ["React", "TypeScript", "Python", "Neo4j", "pgvector", "Ollama", "Docker"],
        "highlights": ["6 products in one suite", "SelfAware Continuum: Unicharm + Neo4j + pgvector + Ollama"],
        "milestones": [
            {"name": "SelfAware Continuum — HVAC intelligence", "status": "Active", "progress": 55, "desc": "Neo4j + pgvector + Ollama"},
            {"name": "AuditShield — Compliance Ledger", "status": "Done", "progress": 100, "desc": "Compliance ledger"},
            {"name": "Autonomous-Cortex — Mission Console", "status": "Done", "progress": 100, "desc": "Mission console"},
            {"name": "NeuralPulse — Signal Lab", "status": "Done", "progress": 100, "desc": "Signal lab"},
            {"name": "VisioScan — Light Studio", "status": "Done", "progress": 100, "desc": "Light studio"},
            {"name": "SpatialNexus — Topology Atlas", "status": "Done", "progress": 100, "desc": "Topology atlas"},
        ],
    },
    # --- REMOVED (2026-06-16, per Harshan) — not in the curated workspace; commented out, restorable ---
    # {
    #     "id": "concept-standard", "name": "Concept Standard — Graph Analytics",
    #     "dir": "concept-standard", "git": False,
    #     "category": "SelfAware Suite", "status": "Active", "progress": 55,
    #     "description": "Interactive concept-map viewer (1,159 concepts, 5,985 relations) for lesson generation.",
    #     "role": "Built the in-browser graph analytics tool (ECharts + Neo4j GDS).",
    #     "tech": ["Python", "JavaScript", "ECharts", "Neo4j"],
    #     "highlights": ["1,159 concepts / 5,985 relations", "Force/circular graphs, command palette, light/dark, a11y"],
    #     "milestones": [],
    # },
    # --- REMOVED (2026-06-16, per Harshan) — not in the curated workspace; commented out, restorable ---
    # {
    #     "id": "voxara", "name": "VOXARA AI — Voice Intelligence",
    #     "dir": "VOXARA AI", "git": False,
    #     "category": "Voice AI", "status": "Paused", "progress": 30,
    #     "description": "Realtime on-prem AI voice intelligence platform — monorepo with local LLM, Whisper STT, Piper TTS.",
    #     "role": "Architecture & monorepo setup (pnpm/Turbo), service scaffolding.",
    #     "tech": ["Node.js", "pnpm", "Turbo", "Ollama", "Whisper", "Piper", "Docker"],
    #     "highlights": ["Largest by file count (37 items)", "Local LLM + Whisper + Piper TTS"],
    #     "milestones": [],
    # },
    # --- REMOVED (2026-06-16, per Harshan) — RANGSONS client app, not Graylinx; commented out, restorable ---
    # {
    #     "id": "travel-app", "name": "TripTrail — Travel & Expense (RANGSONS)",
    #     "dir": "travel-app", "git": True,
    #     "category": "Client Apps", "status": "Done", "progress": 100,
    #     "description": "Buildless travel/expense reporting web app for RANGSONS LLP with HR approval + Google Sheets.",
    #     "role": "Sole developer — static web app, Supabase backend, approval flow.",
    #     "tech": ["HTML", "Tailwind", "JavaScript", "Supabase"],
    #     "highlights": ["HR approval + Google Sheets integration", "Buildless static deployment"],
    #     "milestones": [],
    # },
    # --- REMOVED (2026-06-16, per Harshan) — not in the curated workspace; commented out, restorable ---
    # {
    #     "id": "konva", "name": "Konva Canvas App",
    #     "dir": "konva", "git": False,
    #     "category": "Client Apps", "status": "Paused", "progress": 25,
    #     "description": "React + TypeScript canvas/drawing app template (likely a UI/design component base).",
    #     "role": "Template/prototype setup.",
    #     "tech": ["React", "TypeScript", "Vite", "Konva"],
    #     "highlights": ["Canvas/drawing UI base"],
    #     "milestones": [],
    # },
    {
        "id": "farmer-app", "name": "Nesso — Farm Traceability Platform",
        "dir": "farmer-app", "git": True,
        "category": "Client Apps", "status": "Active", "progress": 45,
        "description": "Farm-to-fork traceability platform for NR Group (Nesso).",
        "role": "Built the traceability web app — including a weekend build sprint (30–31 May).",
        "tech": ["TypeScript", "React", "Node.js"],
        "highlights": ["Farm-to-fork traceability (NR Group)", "~1-week build incl. weekend sprint"],
        "milestones": [],
    },
    {
        "id": "selfaware-dev-stack", "name": "SelfAware Dev Stack",
        "dir": "selfaware-dev-stack", "git": False,
        "category": "Infra & Tooling", "status": "Done", "progress": 100,
        "description": "Canonical local Docker dev stack (MySQL, PostgreSQL, Neo4j, Redis) for the SelfAware products.",
        "role": "Authored the docker-compose environment for 4 products.",
        "tech": ["Docker", "MySQL", "PostgreSQL", "Neo4j", "Redis"],
        "highlights": ["One dev environment for 4 products"],
        "milestones": [],
    },
    {
        "id": "mysql-unicharm-proxy", "name": "MySQL Unicharm Proxy (deprecated)",
        "dir": "mysql-unicharm-proxy", "git": False,
        "category": "Infra & Tooling", "status": "Done", "progress": 100,
        "description": "Docker MySQL port-forwarding helper — superseded by selfaware-dev-stack.",
        "role": "Quick infra utility (now reference-only).",
        "tech": ["Docker", "MySQL"],
        "highlights": ["Deprecated — replaced by selfaware-dev-stack"],
        "milestones": [],
    },
    # --- REMOVED (2026-06-16, per Harshan) — not in the curated workspace; infra, not really "work"; restorable ---
    # {
    #     "id": "kimo-ssh-keys", "name": "Kimo SSH Keys",
    #     "dir": "kimo-ssh-keys", "git": False,
    #     "category": "Infra & Tooling", "status": "Done", "progress": 100,
    #     "description": "SSH key storage / Graylinx laptop identity (kept out of git for security).",
    #     "role": "Access/identity setup.",
    #     "tech": ["SSH"],
    #     "highlights": ["Laptop identity & access keys"],
    #     "milestones": [],
    # },
    {
        "id": "python-etl", "name": "Python ETL",
        "dir": "python-etl", "git": False,
        "category": "Infra & Tooling", "status": "Planned", "progress": 10,
        "description": "ETL / transformer-layer workstream (planning + follow-ups in May).",
        "role": "ETL transformer-layer planning and conversion-workflow design.",
        "tech": ["Python", "ETL"],
        "highlights": ["ETL transformer layer planning"],
        "override": {"start": "2026-05-14", "end": "2026-05-25", "source": "manual"},
        "milestones": [],
    },
    {
        "id": "raspberry", "name": "Raspberry Stuffs",
        "dir": "Rasberry-stuffs", "git": False,
        "category": "Infra & Tooling", "status": "Planned", "progress": 5,
        "description": "Raspberry Pi / edge experiments (placeholder).",
        "role": "Edge hardware experiments.",
        "tech": ["Raspberry Pi"],
        "highlights": ["Placeholder"],
        "milestones": [],
    },
    # --- GitHub-confirmed Graylinx repos (authoritative dates via `gh`, account Kimosabey) ---
    {
        "id": "gl-hvac-early", "name": "Graylinx.ai HVAC (early prototype)",
        "dir": None, "git": False,
        "category": "HVAC Platform", "status": "Done", "progress": 100,
        "description": "\"HVAC Data Pipeline v2\" — first Graylinx HVAC repo (Python), precursor to ClimaStream.",
        "role": "Initial HVAC data-pipeline prototype.",
        "tech": ["Python"], "commits": 8,
        "highlights": ["First HVAC repo on GitHub (8 commits)", "Precursor to ClimaStream"],
        "override": {"start": "2026-04-20", "end": "2026-04-21", "source": "github"},
        "milestones": [],
    },
    {
        "id": "gl-legacy-analysis", "name": "Graylinx Legacy-Code Analysis",
        "dir": None, "git": False,
        "category": "Graylinx Core", "status": "Done", "progress": 100,
        "description": "Audit/analysis of the Graylinx legacy codebase that informed the v2/OMNYX rebuild.",
        "role": "Reviewed & documented legacy code for migration.",
        "tech": ["JavaScript", "Node.js"], "commits": 7,
        "highlights": ["Legacy code audit (7 commits)", "Fed into OMNYX architecture"],
        "override": {"start": "2026-04-28", "end": "2026-05-01", "source": "github"},
        "milestones": [],
    },
    # --- 2025 POC / discovery phase (email-evidenced, pre-implementation) ---
    {
        "id": "poc-hvac-pipeline", "name": "HVAC Data Pipeline (v1 & v2)",
        "dir": None, "git": False,
        "category": "HVAC Platform", "status": "Done", "progress": 100,
        "description": "Graylinx HVAC Chiller-Plant telemetry pipeline: Node.js → MySQL + Kafka → MinIO (Parquet + manifests) → Airflow → ClickHouse → Grafana. v1 + production-ready v2 on the harshan-aiyappa account (Nov 2025).",
        "role": "Built the end-to-end pipeline + its 3 POCs with J. Suryanarayanan & Satish Krishna; milestone-tracked over email.",
        "tech": ["Node.js", "Express", "Kafka", "MinIO", "Parquet", "Airflow", "ClickHouse", "Grafana", "Python", "Docker"],
        "commits": 12,
        "highlights": ["Chiller-plant predictive-maintenance + energy-optimization pipeline", "v1 (19 Nov) + production v2 (26 Nov 2025), harshan-aiyappa account", "POCs: Ingestion · Data Lake · Analytics"],
        "override": {"start": "2025-11-19", "end": "2025-11-26", "source": "github"},
        "milestones": [
            {"name": "POC 1 — Ingestion (Node.js → MySQL + Kafka)", "status": "Done", "progress": 100, "desc": "Simulator → Node.js service → MySQL + Kafka"},
            {"name": "POC 2 — Data Lake (Kafka → MinIO/Parquet)", "status": "Done", "progress": 100, "desc": "Kafka → connector → MinIO Parquet storage"},
            {"name": "POC 3 — Analytics (Airflow → ClickHouse → Grafana)", "status": "Done", "progress": 100, "desc": "MinIO → Airflow → ClickHouse → Grafana dashboards"},
        ],
    },
    # --- COMMENTED OUT (2026-06-16) pending verification ---------------------
    # "RMS Configuration" traces back to an email-evidenced works list Harshan
    # shared from his Gmail, but the description (Vyoda/Tanu/Nagesh teams) and the
    # single Oct-28-2025 date were unverified placeholders. Removed from the tracker
    # for now; restore once the real description/dates/collaborators are confirmed.
    # {
    #     "id": "rms-config", "name": "RMS Configuration",
    #     "dir": None, "git": False,
    #     "category": "Infra & Tooling", "status": "Done", "progress": 100,
    #     "description": "RMS configuration coordination across the Vyoda/Tanu and Nagesh teams.",
    #     "role": "Coordinated RMS configuration.",
    #     "tech": ["Config"],
    #     "highlights": ["RMS configuration", "Oct 2025"],
    #     "override": {"start": "2025-10-28", "end": "2025-10-28", "source": "email"},
    #     "milestones": [],
    # },
]

# Ownership / scope tag per project. Default is "Graylinx"; only the non-Graylinx
# ids are listed here. "Client" = built for another company; "Personal" = Harshan's
# own/independent products. Surfaced as a badge + filter in the dashboard.
SCOPE_BY_ID = {
    # client work (other companies)
    "travel-app": "Client",        # RANGSONS LLP
    "farmer-app": "Client",        # NR Group (Nesso)
    "konva": "Client",             # generic canvas template
    # personal / independent products
    "selfaware-suite": "Personal",
    "selfaware-dev-stack": "Personal",
    "voxara": "Personal",
    "concept-standard": "Personal",
    "raspberry": "Personal",
}


def scope_of(pid):
    return SCOPE_BY_ID.get(pid, "Graylinx")


# Engagement -> project cross-link rules (keyword in notes -> project id)
LINK_RULES = [
    (r"\bbacnet\b|device simulation|simulation flow", "gl-pbs"),
    (r"\bprd\b|architecture module|hvac system", "omnyx"),
    (r"\betl\b|transformer", "python-etl"),
    (r"ai demonstration|ai demo|ai demonstrations", "thermynx"),
    (r"interface|workflow architecture", "omnyx"),
    (r"kafka|stress", "current-works"),
    (r"deployment|coordination|office|architecture discussion", "current-works"),
]


# --------------------------------------------------------------------------- #
#  Date gathering
# --------------------------------------------------------------------------- #
def _run(cmd):
    try:
        out = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8",
                             errors="replace", timeout=60)
        return out.stdout
    except Exception:
        return ""


def git_dates(path):
    """Return (start, end, commit_count, set_of_active_iso_dates) for Harshan's commits."""
    if not os.path.isdir(os.path.join(path, ".git")):
        return None
    raw = _run(["git", "-C", path, "log", "--all", "-i", "--author=harshan",
                "--date=short", "--format=%ad"])
    dates = sorted({ln.strip() for ln in raw.splitlines() if ln.strip()})
    if not dates:
        # fall back to all-author history
        raw = _run(["git", "-C", path, "log", "--all", "--date=short", "--format=%ad"])
        dates = sorted({ln.strip() for ln in raw.splitlines() if ln.strip()})
    if not dates:
        return None
    cnt_raw = _run(["git", "-C", path, "rev-list", "--all", "--count", "-i",
                    "--author=harshan"]).strip()
    try:
        count = int(cnt_raw)
    except ValueError:
        count = len(dates)
    return dates[0], dates[-1], count, set(dates)


# --- nested-repo discovery (catches repos like graylinx-v2/omnyx, the SelfAware
#     sub-repos, farmer-app/nesso-...) so their real git dates/commits + weekend
#     activity are captured, not just top-level repos. ----------------------------
_REPO_CACHE = {}
_ALL_REPOS = None


def discover_repos(max_depth=3):
    """All git repos under BASE (depth<=max_depth), excluding the tracker repo (OUT_DIR)."""
    repos = []
    skip = {"node_modules", ".next", "dist", "build", "__pycache__", ".venv", "venv"}
    base_depth = BASE.rstrip(os.sep).count(os.sep)
    for root, dirs, files in os.walk(BASE):
        if root.rstrip(os.sep).count(os.sep) - base_depth > max_depth:
            dirs[:] = []
            continue
        if ".git" in dirs and os.path.abspath(root) != os.path.abspath(OUT_DIR):
            repos.append(root)
        dirs[:] = [d for d in dirs if d not in skip and d != ".git"]
    return repos


def all_repos():
    global _ALL_REPOS
    if _ALL_REPOS is None:
        _ALL_REPOS = discover_repos()
    return _ALL_REPOS


def repo_dates(repo):
    """(commit_count, set_of_iso_dates) of Harshan's commits in one repo (cached)."""
    if repo in _REPO_CACHE:
        return _REPO_CACHE[repo]
    raw = _run(["git", "-C", repo, "log", "--all", "-i", "--author=harshan",
                "--date=short", "--format=%ad"])
    dates = {ln.strip() for ln in raw.splitlines() if ln.strip() and ln.strip()[0].isdigit()}
    cnt = _run(["git", "-C", repo, "rev-list", "--all", "--count", "-i", "--author=harshan"]).strip()
    try:
        count = int(cnt)
    except ValueError:
        count = len(dates)
    _REPO_CACHE[repo] = (count, dates)
    return _REPO_CACHE[repo]


def git_for_project(project_dir):
    """Aggregate Harshan's git history across a project dir and any nested repos.
    Returns (start, end, total_commits, set_of_dates) or None."""
    if not project_dir:
        return None
    pd = os.path.abspath(project_dir)
    dates, total = set(), 0
    for repo in all_repos():
        rp = os.path.abspath(repo)
        if rp == pd or rp.startswith(pd + os.sep):
            c, ds = repo_dates(repo)
            dates |= ds
            total += c
    if not dates:
        return None
    sd = sorted(dates)
    return sd[0], sd[-1], total, dates


_DAY_COMMITS = None


def repo_commits(repo):
    """[(iso_date, subject)] for Harshan's commits in one repo."""
    raw = _run(["git", "-C", repo, "log", "--all", "-i", "--author=harshan",
                "--date=short", "--format=%ad\x1f%s"])
    out = []
    for ln in raw.splitlines():
        if "\x1f" in ln:
            d, s = ln.split("\x1f", 1)
            d = d.strip()
            if d[:1].isdigit():
                out.append((d, s.strip()))
    return out


def day_commits():
    """date -> {project_id: [commit subjects]} across all repos (cached)."""
    global _DAY_COMMITS
    if _DAY_COMMITS is not None:
        return _DAY_COMMITS
    proj_dirs = sorted(
        [(os.path.abspath(os.path.join(BASE, p["dir"])), p["id"]) for p in PROJECTS if p.get("dir")],
        key=lambda t: len(t[0]), reverse=True)
    dc = {}
    for repo in all_repos():
        rp = os.path.abspath(repo)
        pid = next((pid for d, pid in proj_dirs if rp == d or rp.startswith(d + os.sep)), "")
        if not pid:
            continue
        for ds, subj in repo_commits(repo):
            dc.setdefault(ds, {}).setdefault(pid, []).append(subj)
    _DAY_COMMITS = dc
    return dc


def _short_name(pid):
    for p in PROJECTS:
        if p["id"] == pid:
            return p["name"].split(" —")[0].split(" /")[0].strip()
    return pid


def describe_day(dmap):
    """Descriptive note from a day's {pid:[subjects]} commit map — real git subjects."""
    parts = []
    for pid, subs in sorted(dmap.items(), key=lambda kv: -len(kv[1])):
        clean = [s[:80] for s in subs if s]
        top = "; ".join(clean[:3])
        more = f" (+{len(clean) - 3} more)" if len(clean) > 3 else ""
        parts.append(f"{_short_name(pid)} [{len(subs)}]: {top}{more}")
    return " · ".join(parts)


def mtime_dates(path, max_depth=2):
    """Estimate (start, end) from oldest/newest file mtime, ignoring noise dirs."""
    if not os.path.isdir(path):
        return None
    skip = {"node_modules", ".git", ".next", "dist", "build", "__pycache__", ".venv", "venv"}
    oldest = newest = os.path.getmtime(path)
    base_depth = path.rstrip(os.sep).count(os.sep)
    for root, dirs, files in os.walk(path):
        dirs[:] = [d for d in dirs if d not in skip]
        if root.count(os.sep) - base_depth > max_depth:
            dirs[:] = []
            continue
        for f in files:
            try:
                m = os.path.getmtime(os.path.join(root, f))
                oldest = min(oldest, m)
                newest = max(newest, m)
            except OSError:
                pass
    return (date.fromtimestamp(oldest).isoformat(), date.fromtimestamp(newest).isoformat())


def link_project(notes):
    n = (notes or "").lower()
    for pattern, pid in LINK_RULES:
        if re.search(pattern, n):
            return pid
    return ""


# --------------------------------------------------------------------------- #
#  Build project rows
# --------------------------------------------------------------------------- #
def build_projects():
    rows = []
    for p in PROJECTS:
        path = os.path.join(BASE, p["dir"]) if p.get("dir") else None
        start = end = ""
        source = "approx"
        commits = 0
        active = set()

        if p.get("override"):
            start = p["override"]["start"]
            end = p["override"]["end"]
            source = p["override"].get("source", "manual")
            commits = p.get("commits", 0)   # catalogue may pin a commit count (e.g. GitHub-sourced)
        else:
            g = git_for_project(path)   # auto-detect, incl. nested repos
            if g:
                start, end, commits, active = g[0], g[1], g[2], g[3]
                source = "git"
        if not start and path:
            md = mtime_dates(path)
            if md:
                start, end, source = md[0], md[1], "approx"

        # ongoing if last activity is recent and status is Active/Ongoing
        end_display = end
        if p["status"] in ("Active", "Ongoing") and end and end >= (TODAY - timedelta(days=10)).isoformat():
            end_display = "ongoing"

        row = {
            "id": p["id"], "parent_id": "", "type": "project",
            "name": p["name"], "category": p["category"], "scope": scope_of(p["id"]),
            "description": p["description"], "role": p["role"],
            "tech": p["tech"], "start": start, "end": end_display,
            "date_source": source, "status": p["status"], "progress": p["progress"],
            "commits": commits, "location": (path or ""),
            "highlights": p["highlights"], "links": "", "notes": "",
        }
        rows.append(row)

        # milestones inherit the parent window
        ms = p.get("milestones") or []
        n = len(ms)
        for i, m in enumerate(ms):
            mstart, mend = start, (end if end else "")
            if start and end and n > 1:
                # spread milestones across the parent window
                try:
                    s = datetime.fromisoformat(start).date()
                    e = datetime.fromisoformat(end).date()
                    span = (e - s).days
                    mstart = (s + timedelta(days=int(span * i / n))).isoformat()
                    mend = (s + timedelta(days=int(span * (i + 1) / n))).isoformat()
                except ValueError:
                    pass
            rows.append({
                "id": f"{p['id']}.{i+1}", "parent_id": p["id"], "type": "milestone",
                "name": m["name"], "category": p["category"], "scope": scope_of(p["id"]),
                "description": m.get("desc", ""), "role": "",
                "tech": [], "start": mstart, "end": mend,
                "date_source": source, "status": m["status"], "progress": m["progress"],
                "commits": 0, "location": "", "highlights": [], "links": "", "notes": "",
            })
    return rows


# --------------------------------------------------------------------------- #
#  Build engagement rows
# --------------------------------------------------------------------------- #
def parse_attendance():
    rows = []
    if not os.path.exists(ATTENDANCE_CSV):
        return rows
    with open(ATTENDANCE_CSV, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        data = [r for r in reader if any((c or "").strip() for c in r)]
    if not data:
        return rows
    # header has a leading empty column; columns 1..10 are the real fields
    for r in data[1:]:
        cells = [c.strip() for c in r]
        # pad
        while len(cells) < 11:
            cells.append("")
        _, d, day, wtype, loc, stayed, hotel, cin, cout, travel, notes = cells[:11]
        if not d:
            continue
        try:
            iso = datetime.strptime(d, "%d-%b-%y").date().isoformat()
        except ValueError:
            continue
        rows.append({
            "date": iso, "day": day, "work_type": wtype or "WFH",
            "location": loc, "hotel": hotel if stayed.lower() == "yes" else "",
            "check_in": cin, "check_out": cout, "travel": travel, "notes": notes,
            "linked_project": link_project(notes), "source": "csv",
        })
    return rows


def carry_forward(att, projects):
    """Fill days after the last attendance row up to TODAY using git activity, else needs-fill."""
    if not att:
        return att
    have = {r["date"] for r in att}
    last = max(datetime.fromisoformat(r["date"]).date() for r in att)

    # per-day commit subjects across ALL repos (incl. nested) -> a real, descriptive
    # note for each derived day (captures weekend work in omnyx/nesso/etc. too).
    dc = day_commits()

    extra = []
    d = last + timedelta(days=1)
    while d <= TODAY:
        iso = d.isoformat()
        if iso not in have:
            dm = dc.get(iso)
            if dm:
                pid = max(dm, key=lambda k: len(dm[k]))
                total = sum(len(v) for v in dm.values())
                extra.append({
                    "date": iso, "day": d.strftime("%A"), "work_type": "WFH",
                    "location": "Mysore (residence)", "hotel": "",
                    "check_in": "", "check_out": "", "travel": "",
                    "notes": describe_day(dm) + f"  ({total} commits)",
                    "linked_project": pid, "source": "derived",
                })
            else:
                extra.append({
                    "date": iso, "day": d.strftime("%A"), "work_type": "",
                    "location": "", "hotel": "", "check_in": "", "check_out": "",
                    "travel": "", "notes": "", "linked_project": "", "source": "needs-fill",
                })
        d += timedelta(days=1)
    return att + extra


# Days with no git commits and no attendance row were spent on the Lingotran task
# and recruitment interviews for RAPL & Examic (per Harshan). Recorded here so they
# aren't blank/needs-fill.
GAP_NOTE = "Lingotran task + recruitment interviews for RAPL and Examic."
GAP_DATES = ["2026-04-23", "2026-04-24", "2026-04-25", "2026-04-26",
             "2026-05-02", "2026-05-03", "2026-05-07", "2026-06-14"]


def apply_gap_fill(eng):
    """Fill known non-coding gap days (Lingotran + RAPL/Examic interviews)."""
    by = {r["date"]: r for r in eng}
    for ds in GAP_DATES:
        dt = datetime.fromisoformat(ds).date()
        row = {"date": ds, "day": dt.strftime("%A"), "work_type": "WFH",
               "location": "Mysore", "hotel": "", "check_in": "", "check_out": "",
               "travel": "", "notes": GAP_NOTE, "linked_project": "", "source": "manual"}
        if ds in by:
            by[ds].update(row)
        else:
            eng.append(row)
    return eng


# --------------------------------------------------------------------------- #
#  Teams 1:1 chat -> per-day communications (topics only; never verbatim)
# --------------------------------------------------------------------------- #
WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
_TS_ABS = re.compile(r"^(\d{2})-(\d{2})(?:-(\d{4}))?\s+\d{1,2}:\d{2}\s*[AP]M$")
_TS_REL = re.compile(r"^(Yesterday|Today|Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\s+\d{1,2}:\d{2}\s*[AP]M$")
_TS_BARE = re.compile(r"^\d{1,2}:\d{2}\s*[AP]M$")
_TEAMS_CHROME = {
    "Chat", "Shared", "Storyline", "Unread", "Channels", "Chats", "Meeting chats",
    "Message List", "Has context menu", "has context menu", "Notes", "Community",
    "Temporarily shown", "Press Ctrl+F to find in this chat",
    "Notes have been created in this chat.",
}
_TOPIC_RULES = [
    (r"kafka|stress|throughput|150x|5000x|load[ -]?test|breaking point", "Kafka stress test"),
    (r"\bgpu\b|vram|jarvis|80\s*gb|a100|h100|cuda|rtx|nvidia", "GPU / VRAM sizing"),
    (r"agent|orchestrat|planner|langgraph|devstral|tool[- ]?call", "Agentic orchestration"),
    (r"\beval\b|evaluat|benchmark|accuracy|metric|ground truth", "Model evaluation"),
    (r"\brag\b|retriev|embedding|vector|chunk|pgvector", "RAG pipeline"),
    (r"telemetry|virtual sens|sensor|\bgat\b|\bgnn\b|graph neural", "Telemetry / sensing"),
    (r"deploy|docker|\bexe\b|on-?prem|netlify|nginx|container", "Deployment"),
    (r"hir(e|ing)|interview|candidate|resume|recruit|onboard|shortlist", "Hiring / interviews"),
    (r"quot|price|cost|budget|invoice|rupees|lakh|payment", "Hardware / cost"),
    (r"bacnet|bacpypes|\bpbs\b|device sim|simulat", "BACnet / simulation"),
    (r"schema|database|\bdb\b|postgres|mysql|timescale|backend|\bapi\b", "Backend / DB"),
    (r"architect|design|diagram|\bflow\b|pipeline|module", "Architecture"),
    (r"meet|call|\bsync\b|standup|discuss|\bplan\b|review", "Planning / sync"),
]
_TOPIC_STOP = set("""the and for with from this that have has will would can could should just like get got
need want also more most there here what when where which who how why okay yeah your you our they them
their image media preview link message reaction harshan aiyappa team teams about into your yours been
will well done good thanks please sure into onto over under than then them code clone added connected""".split())
# also drop colleague names (incl. spelling variants) so they never surface as a "topic"
_TOPIC_STOP |= {w.lower() for name in TEAMS_PEOPLE.values() for w in name.split()} | {"santhosh"}


def _resolve_rel(word, C):
    if word == "Today":
        return C
    if word == "Yesterday":
        return C - timedelta(days=1)
    target = WEEKDAYS.index(word)
    d = C
    for _ in range(7):
        if d.weekday() == target:
            return d
        d -= timedelta(days=1)
    return C


def _derive_topic(texts):
    """Short THEME string for a day's chat text. Never returns verbatim message content."""
    blob = re.sub(r"https?://\S+", " ", " ".join(texts).lower())
    tags = []
    for pat, tag in _TOPIC_RULES:
        if re.search(pat, blob) and tag not in tags:
            tags.append(tag)
            if len(tags) >= 2:
                break
    if tags:
        return ", ".join(tags)
    freq = {}
    for w in re.findall(r"[a-z]{4,}", blob):
        if w in _TOPIC_STOP:
            continue
        freq[w] = freq.get(w, 0) + 1
    top = sorted(freq, key=lambda w: (-freq[w], w))[:2]
    return " ".join(t.title() for t in top) if top else "general chat"


def _is_topic_text(ln, person):
    low = ln.lower()
    if ln.startswith("Begin quote") or low.startswith("http") or low.startswith("url preview"):
        return False
    if low in ("image", "media", "url preview") or ln in ("Harshan Aiyappa", person):
        return False
    if "reaction" in low and len(ln) < 24:
        return False
    if " by " in ln and ("attachment" in low or ln.rstrip().endswith(person) or ln.rstrip().endswith("Aiyappa")):
        return False
    return True


def parse_teams_chat():
    """gl_teams_chat/*.txt -> {iso_date: {person: {count, topic, snippet}}}.
    Each file is one 1:1 chat, so every standalone timestamp line = one message with that person.
    Pure + deterministic (no clock/random); year defaults to 2026; relative times anchor on TODAY."""
    if not os.path.isdir(TEAMS_CHAT_DIR):
        return {}
    out = {}
    for stem in sorted(TEAMS_PEOPLE):
        path = os.path.join(TEAMS_CHAT_DIR, stem + ".txt")
        if not os.path.exists(path):
            continue
        try:
            with open(path, encoding="utf-8", errors="replace") as fh:
                lines = fh.read().splitlines()
        except OSError:
            continue
        if not any(ln.strip() for ln in lines):
            continue
        person = TEAMS_PEOPLE[stem]
        cur, last_dt = None, None
        for raw in lines:
            ln = raw.strip()
            if not ln or ln in _TEAMS_CHROME:
                continue
            cand = None
            m = _TS_ABS.match(ln)
            if m:
                yr = int(m.group(3)) if m.group(3) else 2026
                try:
                    cand = date(yr, int(m.group(2)), int(m.group(1)))
                except ValueError:
                    cand = None
            else:
                mr = _TS_REL.match(ln)
                if mr:
                    cand = _resolve_rel(mr.group(1), TODAY)
                elif _TS_BARE.match(ln):
                    cand = TODAY
            if cand is not None:
                if last_dt and (last_dt - cand).days > 3:  # backward jump = quoted/pasted date
                    continue
                cur, last_dt = cand, cand
                rec = out.setdefault(cur.isoformat(), {}).setdefault(person, {"count": 0, "_t": []})
                rec["count"] += 1
                continue
            if cur and _is_topic_text(ln, person):
                out[cur.isoformat()][person]["_t"].append(ln)
    res = {}
    for iso, pm in out.items():
        res[iso] = {}
        for person, rec in pm.items():
            texts = rec["_t"]
            snip = next((t for t in texts if len(t.split()) >= 4), texts[0] if texts else "")
            res[iso][person] = {"count": rec["count"], "topic": _derive_topic(texts),
                                "snippet": snip[:80].rstrip()}
    return res


def _fmt_comm(day_map, detail):
    parts = []
    for person in sorted(day_map):
        rec = day_map[person]
        seg = f"{person} ({rec['count']})"
        if detail in ("topics", "full") and rec.get("topic"):
            seg += f" — {rec['topic']}"
        if detail == "full" and rec.get("snippet"):
            seg += f": “{rec['snippet']}”"
        parts.append(seg)
    return " · ".join(parts)


def apply_communications(eng, comm_by_date, detail=TEAMS_DETAIL):
    """Enrich EXISTING engagement days with a per-day communications summary. Chat dates with no
    matching engagement day are skipped (returns counts so main() can report them)."""
    by = {r["date"]: r for r in eng}
    matched = skipped = 0
    for ds in sorted(comm_by_date):
        if ds in by:
            by[ds]["communications"] = _fmt_comm(comm_by_date[ds], detail)
            matched += 1
        else:
            skipped += 1
    for r in eng:
        r.setdefault("communications", "")
    return eng, matched, skipped


def compute_analytics(proj, engagement):
    """Precompute 'how much worked' analytics for the dashboard Analytics view."""
    by_proj = sorted([[p["name"], p["commits"]] for p in proj if p.get("commits")],
                     key=lambda x: -x[1])[:12]
    cat, status, scope = {}, {}, {}
    for p in proj:
        cat[p["category"]] = cat.get(p["category"], 0) + (p["commits"] or 0)
        status[p["status"]] = status.get(p["status"], 0) + 1
        scope[p.get("scope", "Graylinx")] = scope.get(p.get("scope", "Graylinx"), 0) + 1
    wt = {}
    for r in engagement:
        k = r["work_type"] or "—"
        wt[k] = wt.get(k, 0) + 1
    worked = [r for r in engagement if r["work_type"] and r["work_type"] != "Leave"]
    weekend_worked = sum(1 for r in worked if datetime.fromisoformat(r["date"]).weekday() >= 5)
    per_day = {d: sum(len(v) for v in m.values()) for d, m in day_commits().items()}
    busiest = max(per_day.items(), key=lambda kv: kv[1]) if per_day else ("", 0)
    weekly = {}
    for d, c in per_day.items():
        iso = datetime.fromisoformat(d).date().isocalendar()
        key = f"{iso[1]:02d}"  # ISO week number
        weekly[key] = weekly.get(key, 0) + c
    wdates = sorted(r["date"] for r in worked)
    longest = cur = 0
    prev = None
    for ds in wdates:
        dt = datetime.fromisoformat(ds).date()
        cur = cur + 1 if (prev and (dt - prev).days == 1) else 1
        longest = max(longest, cur)
        prev = dt
    active = len(per_day)
    total = sum(p["commits"] or 0 for p in proj)
    return {
        "per_day": sorted(per_day.items()),
        "commits_by_project": by_proj,
        "commits_by_category": sorted(cat.items(), key=lambda x: -x[1]),
        "projects_by_status": sorted(status.items(), key=lambda x: -x[1]),
        "projects_by_scope": sorted(scope.items(), key=lambda x: -x[1]),
        "work_type": sorted(wt.items(), key=lambda x: -x[1]),
        "weekly_commits": [[f"W{k}", v] for k, v in sorted(weekly.items())],
        "weekend_worked": weekend_worked,
        "worked_days": len(worked),
        "active_commit_days": active,
        "busiest_day": {"date": busiest[0], "commits": busiest[1]},
        "commits_per_active_day": round(total / active, 1) if active else 0,
        "longest_streak": longest,
        "hotel_nights": sum(1 for r in engagement if r["hotel"]),
        "locations": len(set(r["location"] for r in engagement if r["location"])),
    }


# --------------------------------------------------------------------------- #
#  Writers
# --------------------------------------------------------------------------- #
PROJ_COLS = ["id", "parent_id", "type", "name", "category", "scope", "description", "role",
             "tech", "start", "end", "date_source", "status", "progress", "commits",
             "location", "highlights", "links", "notes"]
ENG_COLS = ["date", "day", "work_type", "location", "hotel", "check_in", "check_out",
            "travel", "notes", "linked_project", "source", "communications"]


def _flat(v):
    if isinstance(v, list):
        return "; ".join(str(x) for x in v)
    return v


def write_csv(path, cols, rows):
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(cols)
        for r in rows:
            w.writerow([_flat(r.get(c, "")) for c in cols])


def write_json(path, payload):
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)


def write_xlsx(path, projects, engagement):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ! openpyxl not installed — skipping works.xlsx")
        return

    def fill(hexv):
        return PatternFill("solid", fgColor=hexv)

    thin = Side(style="thin", color=C["line"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_font = Font(bold=True, color="FFFFFF", size=11)
    head_fill = fill(C["ink"])

    wb = Workbook()
    # deterministic metadata so the .xlsx bytes are stable across rebuilds (no git churn)
    stamp = datetime(TODAY.year, TODAY.month, TODAY.day)
    wb.properties.created = stamp
    wb.properties.modified = stamp
    wb.properties.creator = "build_tracker.py"

    def make_sheet(title, cols, rows, status_key=None):
        ws = wb.create_sheet(title)
        ws.append(cols)
        for ci, _ in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center", horizontal="left")
            cell.border = border
        for r in rows:
            ws.append([_flat(r.get(c, "")) for c in cols])
        # status colouring
        if status_key and status_key in cols:
            sidx = cols.index(status_key) + 1
            for ri in range(2, ws.max_row + 1):
                val = ws.cell(row=ri, column=sidx).value
                col = C.get(val)
                if col:
                    for ci in range(1, len(cols) + 1):
                        ws.cell(row=ri, column=ci).fill = fill(_tint(col))
                    sc = ws.cell(row=ri, column=sidx)
                    sc.font = Font(bold=True, color=col)
        # borders + wrap + widths
        for ri in range(2, ws.max_row + 1):
            for ci in range(1, len(cols) + 1):
                cc = ws.cell(row=ri, column=ci)
                cc.border = border
                cc.alignment = Alignment(vertical="top", wrap_text=True)
        for ci, name in enumerate(cols, 1):
            maxlen = max([len(str(name))] + [len(str(_flat(r.get(name, "")))) for r in rows] + [0])
            ws.column_dimensions[get_column_letter(ci)].width = min(max(12, maxlen + 2), 55)
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        return ws

    make_sheet("Projects", PROJ_COLS, projects, status_key="status")
    make_sheet("Engagement", ENG_COLS, engagement, status_key=None)
    # color engagement by work_type
    ws = wb["Engagement"]
    wt_color = {"Office": C["Office"], "Discussion": C["Discussion"], "WFH": C["WFH"],
                "Leave": C["Leave"]}
    wtidx = ENG_COLS.index("work_type") + 1
    sridx = ENG_COLS.index("source") + 1
    for ri in range(2, ws.max_row + 1):
        wt = ws.cell(row=ri, column=wtidx).value
        src = ws.cell(row=ri, column=sridx).value
        if src == "needs-fill":
            ws.cell(row=ri, column=wtidx).fill = fill("E2E2E2")
        col = wt_color.get(wt)
        if col:
            ws.cell(row=ri, column=wtidx).font = Font(bold=True, color=col)

    wb.remove(wb["Sheet"])
    wb.save(path)
    _normalize_xlsx(path)


def _normalize_xlsx(path):
    """Rewrite the .xlsx zip with fixed member timestamps + sorted order so the
    bytes are reproducible across rebuilds (avoids meaningless git churn)."""
    import zipfile
    fixed = (2026, 6, 15, 0, 0, 0)
    with zipfile.ZipFile(path) as zin:
        items = sorted(zin.infolist(), key=lambda i: i.filename)
        blobs = [(i.filename, i.external_attr, zin.read(i.filename)) for i in items]
    tmp = path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, attr, data in blobs:
            if name == "docProps/core.xml":   # openpyxl writes real save-time here
                data = re.sub(rb"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z", b"2026-06-15T00:00:00Z", data)
            zi = zipfile.ZipInfo(name, fixed)
            zi.compress_type = zipfile.ZIP_DEFLATED
            zi.external_attr = attr
            zout.writestr(zi, data)
    os.replace(tmp, path)


def _tint(hexv, amount=0.86):
    """Lighten a hex color toward white for soft row fills."""
    r = int(hexv[0:2], 16); g = int(hexv[2:4], 16); b = int(hexv[4:6], 16)
    r = int(r + (255 - r) * amount); g = int(g + (255 - g) * amount); b = int(b + (255 - b) * amount)
    return f"{r:02X}{g:02X}{b:02X}"


def inject_seed(path, payload):
    if not os.path.exists(path):
        print(f"  ! {os.path.basename(path)} not found — skipping seed injection")
        return
    with open(path, encoding="utf-8") as fh:
        html = fh.read()
    blob = json.dumps(payload, ensure_ascii=False)
    pattern = re.compile(
        r'(<script id="seed-data" type="application/json">).*?(</script>)', re.DOTALL)
    if not pattern.search(html):
        print("  ! seed-data <script> block not found in index.html")
        return
    html = pattern.sub(lambda m: m.group(1) + blob + m.group(2), html, count=1)
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(html)


# Files copied into public/ for the Netlify deploy (self-contained, seed embedded).
PUBLIC_FILES = ["index.html",
                "assets/favicon.svg", "assets/graylinxLogo.png",
                "assets/IMG-20230616-WA0000.jpg", "assets/SpaceGrotesk.woff2",
                "vendor/lenis.min.js", "vendor/echarts.min.js",
                "vendor/gsap.min.js", "vendor/ScrollTrigger.min.js",
                "vendor/three.r134.min.js", "vendor/vanta.net.min.js"]


def build_public():
    """Assemble a clean static deploy dir: dashboard + lenis + images only.

    Excludes source (.py), raw datasets and the personal attendance CSV. The
    dashboard's data is embedded as seed JSON, so no data files are needed at runtime.
    """
    pub = os.path.join(OUT_DIR, "public")
    os.makedirs(pub, exist_ok=True)
    copied = []
    for name in PUBLIC_FILES:
        src = os.path.join(OUT_DIR, name)
        if os.path.exists(src):
            dst = os.path.join(pub, name)
            os.makedirs(os.path.dirname(dst), exist_ok=True)
            shutil.copy2(src, dst)
            copied.append(name)
    return copied


# --------------------------------------------------------------------------- #
#  Main
# --------------------------------------------------------------------------- #
def main():
    print("Building Graylinx x Harshan work tracker ...")
    projects = build_projects()
    engagement = parse_attendance()
    engagement = carry_forward(engagement, projects)
    engagement = apply_gap_fill(engagement)
    engagement, _comm_matched, _comm_skipped = apply_communications(engagement, parse_teams_chat())
    engagement.sort(key=lambda r: r["date"])

    proj_only = [r for r in projects if r["type"] == "project"]
    span_start = min([r["start"] for r in proj_only if r["start"]] +
                     [r["date"] for r in engagement] or [TODAY.isoformat()])
    on_site = sum(1 for r in engagement if r["work_type"] in ("Office", "Discussion"))
    wfh = sum(1 for r in engagement if r["work_type"] == "WFH")
    hotel_nights = sum(1 for r in engagement if r["hotel"])
    needs_fill = sum(1 for r in engagement if r["source"] == "needs-fill")
    total_commits = sum(r["commits"] for r in proj_only)

    meta = {
        "title": "Harshan × Graylinx — Work Tracker",
        "generated": TODAY.isoformat(),
        "span_start": span_start,
        "span_end": TODAY.isoformat(),
        "stats": {
            "projects": len(proj_only),
            "active": sum(1 for r in proj_only if r["status"] in ("Active", "Ongoing")),
            "done": sum(1 for r in proj_only if r["status"] == "Done"),
            "total_commits": total_commits,
            "engagement_days": len(engagement),
            "on_site_days": on_site,
            "wfh_days": wfh,
            "hotel_nights": hotel_nights,
            "needs_fill": needs_fill,
        },
        "categories": sorted({r["category"] for r in proj_only}),
        "palette": C,
    }
    meta["analytics"] = compute_analytics(proj_only, engagement)
    # Content signature: when this changes, the dashboard auto-reloads the fresh
    # seed (so a stale localStorage copy never masks regenerated data).
    meta["build"] = hashlib.sha1(
        (json.dumps(projects, sort_keys=True) + json.dumps(engagement, sort_keys=True)).encode("utf-8")
    ).hexdigest()[:12]
    payload = {"meta": meta, "projects": projects, "engagement": engagement}

    os.makedirs(os.path.join(OUT_DIR, "data"), exist_ok=True)
    write_csv(os.path.join(OUT_DIR, "data", "works.csv"), PROJ_COLS, projects)
    write_csv(os.path.join(OUT_DIR, "data", "engagement.csv"), ENG_COLS, engagement)
    write_json(os.path.join(OUT_DIR, "data", "works.json"), payload)
    write_xlsx(os.path.join(OUT_DIR, "data", "works.xlsx"), projects, engagement)
    inject_seed(INDEX_HTML, payload)
    pub = build_public()

    print(f"  projects: {len(proj_only)} (+{len(projects)-len(proj_only)} milestones)")
    print(f"  engagement days: {len(engagement)} (needs-fill: {needs_fill})")
    print(f"  teams chat: detail={TEAMS_DETAIL} | {_comm_matched} days enriched, {_comm_skipped} out-of-span skipped")
    print(f"  span: {span_start} -> {TODAY.isoformat()} | commits: {total_commits}")
    print("  wrote works.csv, engagement.csv, works.json, works.xlsx; injected seed into index.html")
    print(f"  public/ deploy dir: {', '.join(pub) if pub else '(no files copied)'}")
    print("Done.")


if __name__ == "__main__":
    main()
